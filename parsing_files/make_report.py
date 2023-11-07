'''Задача: написать функцию make_report,
которая парсит журнал логов (файл logs.xlsx) и создает отчет по
самым популярным браузерам и самым продаваемым товарам.'''

import pandas
from pprint import pprint
import collections
from openpyxl import load_workbook


def make_report(log_file_name, report_template_file_name, report_output_file_name):
    # считываем данные из файла
    excel_data = pandas.read_excel(log_file_name, sheet_name='log', engine='openpyxl')

    # преобразуем данные к словарю
    excel_data_dict = excel_data.to_dict(orient='records')

    # создаю список, добавляю туда все браузеры из документа
    list_browser = []
    for i in range(len(excel_data_dict)):
        list_browser.append(excel_data_dict[i]['Браузер'])

    # подсчитываем количество посещений среди всех браузеров
    browser_counter = collections.Counter(list_browser)
    # список 7ми самых популярных браузеров
    best_browser = browser_counter.most_common(7)

    # создаю словарь, чтобы найти все браузеры(ключ) = [{количество их посещений по месяцам(значение)}]
    dict_month = {}
    for i in range(len(excel_data_dict)):
        dict_month[excel_data_dict[i]['Браузер']] = [{
            1: 0,
            2: 0,
            3: 0,
            4: 0,
            5: 0,
            6: 0,
            7: 0,
            8: 0,
            9: 0,
            10: 0,
            11: 0,
            12: 0}]
    # записываю данные в словарь dict_month
    for i in range(len(excel_data_dict)):  # прохожусь по заданному документу
        browser = excel_data_dict[i]['Браузер']  # название браузера
        date_month = excel_data_dict[i]['Дата посещения'].month  # месяц его посещения
        # прибавляем +1 в словарь по имени браузера и месяцу посещений
        dict_month[browser][0][date_month] = dict_month[browser][0][date_month] + 1

    # создаю словарь, чтобы найти все товары(ключ) и количество их заказов(значение)
    dict_product = {}
    # создаю список, со всеми товарами из документа
    list_product = []
    for i in range(len(excel_data_dict)):
        # проходимся по списку товаров каждого клиента
        for product in excel_data_dict[i]['Купленные товары'].split(','):
            # ключ - продукт; к значению +1 если встретили такой ключ
            dict_product[product] = dict_product.get(product, 0) + 1
            # список всех товаров
            list_product.append(product)

    # подсчитываем количество всех товаров
    product_counter = collections.Counter(list_product)
    # список самых популярных 7 товаров
    best_product = product_counter.most_common(7)

    # создаем словарь, чтобы вычислить предпочтения по полу людей
    dict_sex = {}
    # ключ - товар, значение - список со словарем, где ключи муж и жен
    for product in dict_product.keys():
        dict_sex[product] = [{'м': 0,
                              'ж': 0}]
    # записываю данные в словарь dict_sex
    for i in range(len(excel_data_dict)):
        # достаем пол заказавшего i-ый список товаров
        sex = excel_data_dict[i]['Пол']
        # проходим по i-тому списку товаров и записываем пол в наш словарь dict_sex
        for products in excel_data_dict[i]['Купленные товары'].split(','):
            dict_sex[products][0][sex] = dict_sex[products][0][sex] + 1

    # создаем переменные для дальнейшего сравнения
    woman_max = -1
    woman_max_product = ''
    woman_min = 10 ** 5
    woman_min_product = ''
    man_max = -1
    man_max_product = ''
    man_min = 10 ** 5
    man_min_product = ''

    # проходимся по словарю dict_sex и ищем максимальные и минимальные заказы среди женщин
    for items in dict_sex.items():
        if items[-1][0]['ж'] > woman_max:
            woman_max = items[-1][0]['ж']
            woman_max_product = items[0]
        elif items[-1][0]['ж'] < woman_min:
            woman_min = items[-1][0]['ж']
            woman_min_product = items[0]

    # проходимся по словарю dict_sex и ищем максимальные и минимальные заказы среди мужчин
    for items in dict_sex.items():
        if items[-1][0]['м'] > man_max:
            man_max = items[-1][0]['м']
            man_max_product = items[0]
        elif items[-1][0]['м'] < man_min:
            man_min = items[-1][0]['м']
            man_min_product = items[0]

    # открываем шаблон для записи данных
    wb = load_workbook(filename=report_template_file_name, data_only=True)
    # делаем активный выбранный лист
    sheet = wb['Лист1']
    # записываем лучшие браузеры в шаблон
    for i in range(len(best_browser)):
        sheet[f'A{5 + i}'] = best_browser[i][0]

    # записываем количество посещений по месяцам
    a, b = 5, 2  # начальные координаты ячеек
    # проходимся по лучшим браузерам
    for browsers in best_browser:
        # создаем список, для добавления количества посещения для данного браузера
        list_month = []
        for month in dict_month[browsers[0]][0].values():
            # записываем список посещений
            list_month.append(month)
        # проходимся по списке и записываем посещения по месяцам в колонки
        for i in range(len(list_month)):
            sheet.cell(row=a, column=b + i).value = list_month[i]
        # делаем переход на строчку ниже
        a += 1

    # записываем лучшие продукты в шаблон
    for i in range(len(best_product)):
        sheet[f'A{19 + i}'] = best_product[i][0]

    # записываем популярные и не очень продукты среди муж и жен в шаблон
    sheet[f'B31'] = man_max_product
    sheet[f'B32'] = woman_max_product
    sheet[f'B33'] = man_min_product
    sheet[f'B34'] = woman_min_product

    # сохраняем и закрываем документ
    wb.save(report_output_file_name)


make_report('logs.xlsx', 'report_template.xlsx', 'output_file_name.xlsx')